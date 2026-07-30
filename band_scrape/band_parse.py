# -*- coding: utf-8 -*-
"""
밴드 판매실적 - 2단계 (파싱 + 엑셀 + 사진 다운로드)

사용법 (사장님 PC, band_scrape.py를 돌린 그 폴더에서):
  python3 -m pip install openpyxl requests
  python3 band_parse.py

동작:
  - captured/*.json 을 모두 읽어 사장님 밴드(band_no=82647287) 글만 모음
  - '판매' 실적 글을 골라 모델명·판매일 파싱
  - sales_records.xlsx 로 저장
  - 각 글의 사진을 sales_photos/{판매일}_{모델}_{글번호}/ 폴더로 다운로드
"""
import json, glob, re, os, datetime as dt

BAND_NO = 82647287
CAP_DIR = "captured"


def walk_posts(o, out):
    if isinstance(o, dict):
        if isinstance(o.get('content'), str) and 'post_no' in o:
            out.append(o)
        for v in o.values():
            walk_posts(v, out)
    elif isinstance(o, list):
        for v in o:
            walk_posts(v, out)


def load_band_posts():
    posts = {}
    for f in glob.glob(os.path.join(CAP_DIR, "resp_*.json")):
        try:
            d = json.load(open(f, encoding="utf-8"))
        except Exception:
            continue
        found = []
        walk_posts(d.get("body", {}), found)
        for p in found:
            if (p.get("band", {}) or {}).get("band_no") == BAND_NO:
                posts[p["post_no"]] = p
    return list(posts.values())


MODEL_RE = re.compile(r'\b([A-Z]{2,4}\s?-?\s?\d{3,4}[A-Z]{0,2})\b')

def parse_model(content):
    lines = [l.strip() for l in content.split("\n") if l.strip()]
    # 표준: '판매장비' 다음 줄
    for i, l in enumerate(lines):
        if l.startswith("판매장비") and i + 1 < len(lines):
            m = MODEL_RE.search(lines[i + 1])
            if m:
                return m.group(1).replace(" ", "").replace("-", "-")
    # 폴백: 본문 전체에서 모델 코드
    m = MODEL_RE.search(content)
    return m.group(1).replace(" ", "") if m else ""


def parse_date(content):
    # 라벨 뒤 우선
    m = re.search(r'판매일\s*[:：\-]?\s*(.+)', content)
    seg = m.group(1) if m else content
    for pat, f in [
        (r'(\d{4})\s*[년.]\s*(\d{1,2})\s*[월.]\s*(\d{1,2})', lambda g: (int(g[0]), int(g[1]), int(g[2]))),
        (r'(\d{2})\s*년\s*(\d{1,2})\s*월\s*(\d{1,2})', lambda g: (2000 + int(g[0]), int(g[1]), int(g[2]))),
        (r'\b(20\d{2})(\d{2})(\d{2})\b', lambda g: (int(g[0]), int(g[1]), int(g[2]))),
        (r'(\d{4})\s*년\s*(\d{1,2})\s*월', lambda g: (int(g[0]), int(g[1]), 1)),
    ]:
        mm = re.search(pat, seg)
        if mm:
            try:
                y, mo, da = f(mm.groups())
                return f"{y:04d}-{mo:02d}-{da:02d}"
            except Exception:
                pass
    return ""


def photo_urls(post):
    photos = ((post.get("attachment") or {}).get("photo")) or []
    return [ph.get("photo_url") for ph in photos if ph.get("photo_url")]


def build_rows(posts):
    rows = []
    for p in posts:
        c = p["content"]
        if "판매" not in c:
            continue
        model = parse_model(c)
        sale = parse_date(c)
        standard = c.strip().startswith("판매장비") and bool(model) and bool(sale)
        rows.append({
            "판매일": sale,
            "모델명": model,
            "사진수": p.get("photo_count", 0),
            "작성일": dt.datetime.fromtimestamp(p["created_at"] / 1000).strftime("%Y-%m-%d"),
            "작성자": (p.get("author") or {}).get("name", ""),
            "글번호": p["post_no"],
            "상태": "표준" if standard else "확인필요",
            "원문": c.replace("\n", " / "),
            "링크": p.get("web_url", ""),
            "사진URL": photo_urls(p),
        })
    # 판매일(없으면 작성일) 기준 최신순
    rows.sort(key=lambda r: r["판매일"] or r["작성일"], reverse=True)
    return rows


def write_excel(rows, path="sales_records.xlsx"):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook(); ws = wb.active; ws.title = "판매실적"
    headers = ["판매일", "모델명", "수량(사진)", "작성일", "작성자", "글번호", "상태", "원문", "링크"]
    ws.append(headers)
    for r in rows:
        ws.append([r["판매일"], r["모델명"], r["사진수"], r["작성일"], r["작성자"],
                   r["글번호"], r["상태"], r["원문"], r["링크"]])
    navy = PatternFill("solid", fgColor="1F286F"); tint = PatternFill("solid", fgColor="EEF0F8")
    warn = PatternFill("solid", fgColor="FDE7E3")
    t = Side("thin", color="D0D4E4"); bd = Border(t, t, t, t)
    for c in ws[1]:
        c.font = Font(name="Arial", bold=True, color="FFFFFF"); c.fill = navy
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = bd
    for ri, row in enumerate(ws.iter_rows(min_row=2), 2):
        for c in row:
            c.font = Font(name="Arial", size=10); c.border = bd
            c.alignment = Alignment(vertical="center")
        fill = warn if ws.cell(ri, 7).value == "확인필요" else (tint if ri % 2 == 0 else None)
        if fill:
            for c in row:
                c.fill = fill
    for i, w in enumerate([12, 12, 10, 12, 10, 8, 10, 60, 30], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:I{ws.max_row}"
    wb.save(path)
    return path


def download_photos(rows, base="sales_photos"):
    import requests
    os.makedirs(base, exist_ok=True)
    for r in rows:
        urls = r["사진URL"]
        if not urls:
            continue
        tag = f"{r['판매일'] or r['작성일']}_{r['모델명'] or 'unknown'}_{r['글번호']}"
        d = os.path.join(base, re.sub(r'[^\w.-]', '_', tag))
        os.makedirs(d, exist_ok=True)
        for i, u in enumerate(urls, 1):
            try:
                resp = requests.get(u, timeout=30)
                resp.raise_for_status()
                open(os.path.join(d, f"{i}.jpg"), "wb").write(resp.content)
            except Exception as e:
                print(f"  사진 실패 {tag} #{i}: {e}")
        print(f"  ↓ {tag}  ({len(urls)}장)")


if __name__ == "__main__":
    posts = load_band_posts()
    rows = build_rows(posts)
    n_std = sum(1 for r in rows if r["상태"] == "표준")
    print(f"판매실적 글 {len(rows)}개 (표준 {n_std} / 확인필요 {len(rows)-n_std})")
    p = write_excel(rows)
    print("엑셀 저장:", p)
    print("사진 다운로드 시작...")
    download_photos(rows)
    print("완료 ✅  sales_records.xlsx + sales_photos/ 폴더 확인")
